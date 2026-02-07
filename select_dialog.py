# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'select_dialog.ui'
##
## Created by: Qt User Interface Compiler version 6.10.2
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QDialog, QListWidget, QListWidgetItem,
    QPushButton, QSizePolicy, QWidget)
from openpyxl import load_workbook

class Ui_SelectDialog(object):
    def setupUi(self, SelectDialog):
        if not SelectDialog.objectName():
            SelectDialog.setObjectName(u"SelectDialog")
        SelectDialog.resize(400, 254)
        SelectDialog.setMinimumSize(QSize(400, 254))
        SelectDialog.setMaximumSize(QSize(400, 254))
        self.transList = QListWidget(SelectDialog)
        self.transList.setObjectName(u"transList")
        self.transList.setGeometry(QRect(10, 10, 381, 192))
        self.transList.setStyleSheet(u"font: 12pt \"Calibri\";")
        self.OkBtn = QPushButton(SelectDialog)
        self.OkBtn.setObjectName(u"OkBtn")
        self.OkBtn.setGeometry(QRect(80, 210, 101, 31))
        self.OkBtn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.CancelBtn = QPushButton(SelectDialog)
        self.CancelBtn.setObjectName(u"CancelBtn")
        self.CancelBtn.setGeometry(QRect(220, 210, 101, 31))
        self.CancelBtn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

        self.retranslateUi(SelectDialog)

        QMetaObject.connectSlotsByName(SelectDialog)
    # setupUi

    def retranslateUi(self, SelectDialog):
        SelectDialog.setWindowTitle(QCoreApplication.translate("SelectDialog", u"\u0412\u044b\u0431\u043e\u0440 \u043f\u0435\u0440\u0435\u0432\u043e\u0437\u0447\u0438\u043a\u0430", None))
        self.OkBtn.setText(QCoreApplication.translate("SelectDialog", u"OK", None))
        self.CancelBtn.setText(QCoreApplication.translate("SelectDialog", u"\u041e\u0442\u043c\u0435\u043d\u0430", None))
    # retranslateUi
class SelectDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_SelectDialog()
        self.ui.setupUi(self)

        self.selected_text = None
        self.ui.OkBtn.clicked.connect(self.acceptSelection)
        self.ui.CancelBtn.clicked.connect(self.reject)
        self.load_Trans()
        self.ui.transList.setWordWrap(True)
        self.ui.transList.itemDoubleClicked.connect(self.acceptSelection)

    def acceptSelection(self):
        item = self.ui.transList.currentItem()

        if item:
            self.selected_text = item.text()
            self.accept()

    def load_Trans(self):
        wb = load_workbook("data/template.xltx")
        ws = wb["Данные"]
        row = 2
        while True:
            value = ws[f"C{row}"].value
            if value is None:
                break
            self.ui.transList.addItem(str(value))
            row+=1
