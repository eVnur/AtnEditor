from PySide6.QtWidgets import QApplication, QMainWindow, QDialog, QCompleter, QMessageBox
from PySide6.QtCore import QDate, Qt, QSortFilterProxyModel, QStringListModel
from main_ui import Ui_MainWindow
from openpyxl import load_workbook
from select_dialog import SelectDialog
import subprocess, os, webbrowser,sys, re
import win32com.client

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.load_data()
        self.load_LMP_adress()
        self.ui.transferBtn.clicked.connect(self.open_select_dialog)
        self.ui.PrintBtn.clicked.connect(self.save_to_excel)
        

    def resource_path(self,relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)
    
    def load_data(self):
        wb = load_workbook(self.resource_path("data/template.xltx"))
        ws = wb["titledTN"]
        self.ui.shipperLine.setPlainText(str(ws["B12"].value))
        self.ui.driverLine.setPlainText(str(ws["AD41"].value))
        self.ui.driverName.setText(str(ws["AC65"].value))
        self.ui.mppName.setText(str(ws["C65"].value))
        self.ui.vehicleName.setText(str(ws["B44"].value))
        self.ui.vehicleNumber.setText(str(ws["AD44"].value))
        self.ui.LmpChoise.setCurrentText(str(ws["B55"].value))
        self.ui.dateEdit.setDate(QDate.currentDate())
        self.ui.transLine.setPlainText(str(ws["B41"].value))

    def load_LMP_adress(self):
        wb = load_workbook(self.resource_path("data/template.xltx"))
        ws = wb["Данные"]
        items = []
        row = 2
        while True:
            value = ws[f"A{row}"].value
            if value is None:
                break
            items.append(str(value))
            row += 1
        self.ui.LmpChoise.clear()
        self.ui.LmpChoise.addItems(items)

        model = QStringListModel(items)

        proxy = QSortFilterProxyModel(self)
        proxy.setSourceModel(model)
        proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)
        proxy.setFilterKeyColumn(0)

        completer = QCompleter(proxy, self.ui.LmpChoise)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        def update_filter(text):
            words = text.split()
            if not words:
                regex = ""
            else:
                regex = "|".join([re.escape(word) for word in words])
            proxy.setFilterRegularExpression(regex)
            completer.complete()
        self.ui.LmpChoise.lineEdit().textEdited.connect(update_filter)
        self.ui.LmpChoise.setCompleter(completer)


    def open_select_dialog(self):
        dialog = SelectDialog(self)
        result = dialog.exec()
        if result == 1:
            self.ui.transLine.setPlainText(dialog.selected_text)

    def save_to_excel(self):
        box_value = self.ui.ValueSpinBox.value()
        wb = load_workbook(self.resource_path("data/template.xltx"))
        ws = wb["titledTN"]

        #Данные о грузоперевозчике
        ws["B12"] = self.ui.shipperLine.toPlainText()
        ws["B41"] = self.ui.transLine.toPlainText()

        #Данные водителя
        ws["AD41"] = self.ui.driverLine.toPlainText()
        ws["AC65"] = self.ui.driverName.text()
        ws["AC83"] = self.ui.driverName.text()

        #ФИО МПП
        ws["C65"] = self.ui.mppName.text()

        #Данные о транспорте
        ws["B44"] = self.ui.vehicleName.text()
        ws["AD44"] = self.ui.vehicleNumber.text()

        #Адрес ПВЗ
        ws["B55"] = self.ui.LmpChoise.currentText()
        ws["K8"] = self.ui.LmpChoise.currentText()
        
        #Дата
        ws["D8"] = self.ui.dateEdit.date().toString("dd.MM.yyyy")
        ws["AF8"] = self.ui.dateEdit.date().toString("dd.MM.yyyy")
        ws["Q29"] = self.ui.dateEdit.date().toString("dd.MM.yyyy")

        #Колличество коробок
        ws["P61"] = str(box_value)
        ws["AY22"] = str(box_value)
        ws["AD22"] = f"Количество мест: паллет — {box_value} / пласт.кор - 0,    карт.кор"
        ws["AZ47"] = str(self.ui.comboBox.currentIndex()+1)

        if self.ui.checkBox.isChecked():
            ws["N11"] = '✓'
        else:
            ws["N11"] = ''
        wb.save(self.resource_path("data/template.xltx"))
        self.excel_to_pdf()

    def excel_to_pdf(self):
        BASE_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
        excel_path = self.resource_path("data/template.xltx")
        outdir = self.resource_path("data")

        libreoffice_path = r"C:/Program Files/LibreOffice/program/soffice.exe"

        if os.path.exists(libreoffice_path):
            try:
                subprocess.run([
                    libreoffice_path,
                    "--headless",
                    "--convert-to", "pdf",
                    excel_path,
                    "--outdir", outdir
                ], check=True)

                pdf_file = os.path.join(outdir, "template.pdf")
                webbrowser.open(pdf_file)
                return
            except Exception as e:
                print("Ошибка LibreOffice:", e)

        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(excel_path))
            pdf_file = os.path.join(outdir, "template.pdf")
            wb.ExportAsFixedFormat(0, pdf_file)
            wb.Close(False)
            excel.Quit()
            webbrowser.open(pdf_file)
            return

        except Exception as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText("Не найден способ конвертации в PDF")
            msg.setInformativeText(
                "Установите один из вариантов:\n\n"
                "• LibreOffice\n"
                "• Microsoft Excel\n\n"
                "Без одного из них экспорт невозможен."
            )
            msg.exec()




app = QApplication([])
window = MainWindow()
window.show()
app.exec()
